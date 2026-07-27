"""
Geração do relatório de atendimentos em Excel (.xlsx) — mensal ou
semanal, de um único setor ou de todos os setores juntos (uso
exclusivo da Direção).

O arquivo tem duas abas:

* "Resumo"        — números consolidados para bater o olho: total de
  atendimentos, quantos foram concluídos, cancelados, quantos ficaram
  pendentes, tempo médio de espera e de atendimento, além da quebra por
  assunto, por servidor e — quando o relatório junta vários setores —
  por setor também.
* "Atendimentos"  — uma linha por atendimento, com todos os detalhes
  documentados (cidadão, documentos, assunto, relato, horários, quem
  atendeu, setor, resultado e observações).
"""

import io
from calendar import monthrange
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.atendimento import StatusAtendimento
from app.utils.formatacao import formatar_cpf, formatar_masp


MESES = [
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]

ROTULOS_STATUS = {
    StatusAtendimento.AGUARDANDO.value: "Aguardando",
    StatusAtendimento.CONVOCADO.value: "Convocado",
    StatusAtendimento.EM_ATENDIMENTO.value: "Em atendimento",
    StatusAtendimento.FINALIZADO.value: "Concluído",
    StatusAtendimento.CANCELADO.value: "Cancelado",
}

ROTULOS_ASSUNTO = {
    "documentacao": "Entrega ou consulta de documentação",
    "vida-escolar": "Vida escolar",
    "servidor": "Assuntos relacionados a servidor",
    "outros": "Outros assuntos",
}

COR_CABECALHO = "1E3A5F"
COR_FAIXA = "EEF2F7"

FONTE_CABECALHO = Font(bold=True, color="FFFFFF", size=11)
FONTE_TITULO = Font(bold=True, size=14, color="1E3A5F")
FONTE_SUBTITULO = Font(bold=True, size=11, color="1E3A5F")

BORDA_FINA = Border(
    left=Side(style="thin", color="D6DEE8"),
    right=Side(style="thin", color="D6DEE8"),
    top=Side(style="thin", color="D6DEE8"),
    bottom=Side(style="thin", color="D6DEE8"),
)


def intervalo_do_mes(ano: int, mes: int) -> tuple[datetime, datetime]:
    """Primeiro instante do mês e último instante do mês (inclusive)."""
    if not 1 <= mes <= 12:
        raise ValueError("Mês inválido. Informe um valor entre 1 e 12.")

    inicio = datetime(ano, mes, 1)

    ultimo_dia = monthrange(ano, mes)[1]

    fim = datetime(ano, mes, ultimo_dia, 23, 59, 59, 999999)

    return inicio, fim


def intervalo_da_semana(
    data_referencia: datetime,
) -> tuple[datetime, datetime]:
    """
    Segunda-feira 00:00 até domingo 23:59:59 da semana que contém
    ``data_referencia``. A pessoa escolhe qualquer dia daquela semana
    na tela — não precisa saber o número da semana no calendário.
    """
    inicio_do_dia = data_referencia.replace(
        hour=0, minute=0, second=0, microsecond=0
    )

    segunda_feira = inicio_do_dia - timedelta(
        days=inicio_do_dia.weekday()
    )

    domingo = segunda_feira + timedelta(
        days=6, hours=23, minutes=59, seconds=59, microseconds=999999
    )

    return segunda_feira, domingo


def _rotulo_assunto(assunto: str | None) -> str:
    if not assunto:
        return "Não informado"

    return ROTULOS_ASSUNTO.get(assunto, assunto)


def _formatar_datahora(valor: datetime | None) -> str:
    if not valor:
        return ""

    return valor.strftime("%d/%m/%Y %H:%M")


def _minutos_entre(
    inicio: datetime | None,
    fim: datetime | None,
) -> float | None:
    if not inicio or not fim:
        return None

    return round((fim - inicio).total_seconds() / 60, 1)


def _media(valores: list[float]) -> str:
    if not valores:
        return "—"

    return f"{round(sum(valores) / len(valores), 1)} min".replace(".", ",")


def _escrever_cabecalho(aba, linha: int, colunas: list[str]) -> None:
    for indice, titulo in enumerate(colunas, start=1):
        celula = aba.cell(row=linha, column=indice, value=titulo)
        celula.font = FONTE_CABECALHO
        celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        celula.border = BORDA_FINA


def _ajustar_larguras(aba, larguras: list[int]) -> None:
    for indice, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(indice)].width = largura


def _montar_aba_resumo(
    aba,
    atendimentos: list,
    titulo_periodo: str,
    titulo_abrangencia: str,
    agregado: bool,
) -> None:
    aba.sheet_view.showGridLines = False

    aba["A1"] = "Relatório de atendimentos"
    aba["A1"].font = FONTE_TITULO

    aba["A2"] = f"{titulo_abrangencia} — {titulo_periodo}"
    aba["A2"].font = Font(size=11, color="4A5A6A")

    aba["A3"] = (
        "Emitido em "
        + datetime.now().strftime("%d/%m/%Y às %H:%M")
    )
    aba["A3"].font = Font(size=9, italic=True, color="7A8A9A")

    total = len(atendimentos)

    contagem_status: dict[str, int] = {}
    contagem_assunto: dict[str, int] = {}
    contagem_servidor: dict[str, int] = {}
    contagem_setor: dict[str, int] = {}

    esperas: list[float] = []
    duracoes: list[float] = []
    prioritarios = 0

    for atendimento in atendimentos:
        status = atendimento.status
        contagem_status[status] = contagem_status.get(status, 0) + 1

        assunto = _rotulo_assunto(atendimento.assunto)
        contagem_assunto[assunto] = contagem_assunto.get(assunto, 0) + 1

        if atendimento.servidor_nome:
            nome = atendimento.servidor_nome
            contagem_servidor[nome] = contagem_servidor.get(nome, 0) + 1

        if agregado and atendimento.setor:
            nome_setor = atendimento.setor.nome
            contagem_setor[nome_setor] = (
                contagem_setor.get(nome_setor, 0) + 1
            )

        if atendimento.prioridade == "PRIORITARIO":
            prioritarios += 1

        espera = _minutos_entre(
            atendimento.data_solicitacao,
            atendimento.data_inicio or atendimento.data_convocacao,
        )
        if espera is not None:
            esperas.append(espera)

        duracao = _minutos_entre(
            atendimento.data_inicio,
            atendimento.data_finalizacao,
        )
        if duracao is not None:
            duracoes.append(duracao)

    concluidos = contagem_status.get(
        StatusAtendimento.FINALIZADO.value, 0
    )
    cancelados = contagem_status.get(
        StatusAtendimento.CANCELADO.value, 0
    )
    pendentes = total - concluidos - cancelados

    indicadores = [
        ("Total de atendimentos no período", total),
        ("Concluídos", concluidos),
        ("Cancelados", cancelados),
        ("Ainda em aberto (fila ou em andamento)", pendentes),
        ("Atendimentos prioritários", prioritarios),
        ("Tempo médio de espera na fila", _media(esperas)),
        ("Tempo médio de atendimento", _media(duracoes)),
    ]

    linha = 5
    aba.cell(row=linha, column=1, value="Indicadores do período").font = (
        FONTE_SUBTITULO
    )

    linha += 1
    _escrever_cabecalho(aba, linha, ["Indicador", "Valor"])

    for rotulo, valor in indicadores:
        linha += 1
        aba.cell(row=linha, column=1, value=rotulo).border = BORDA_FINA
        celula_valor = aba.cell(row=linha, column=2, value=valor)
        celula_valor.border = BORDA_FINA
        celula_valor.alignment = Alignment(horizontal="center")

    def _bloco(titulo: str, dados: dict[str, int], inicio: int) -> int:
        aba.cell(row=inicio, column=1, value=titulo).font = FONTE_SUBTITULO

        atual = inicio + 1
        _escrever_cabecalho(aba, atual, ["Descrição", "Quantidade", "%"])

        if not dados:
            atual += 1
            aba.cell(
                row=atual,
                column=1,
                value="Nenhum atendimento no período.",
            ).border = BORDA_FINA
            return atual + 2

        ordenados = sorted(
            dados.items(),
            key=lambda item: item[1],
            reverse=True,
        )

        for descricao, quantidade in ordenados:
            atual += 1
            aba.cell(row=atual, column=1, value=descricao).border = (
                BORDA_FINA
            )

            celula_quantidade = aba.cell(
                row=atual, column=2, value=quantidade
            )
            celula_quantidade.border = BORDA_FINA
            celula_quantidade.alignment = Alignment(horizontal="center")

            percentual = (quantidade / total) if total else 0
            celula_percentual = aba.cell(
                row=atual, column=3, value=percentual
            )
            celula_percentual.number_format = "0.0%"
            celula_percentual.border = BORDA_FINA
            celula_percentual.alignment = Alignment(horizontal="center")

        return atual + 2

    # Quando o relatório junta todos os setores, a quebra por setor vem
    # logo no início — é a primeira pergunta que a Direção costuma ter.
    if agregado:
        linha = _bloco(
            "Atendimentos por setor",
            contagem_setor,
            linha + 2,
        )

    linha = _bloco(
        "Atendimentos por assunto",
        contagem_assunto,
        linha + 2 if not agregado else linha,
    )

    linha = _bloco(
        "Atendimentos por servidor",
        contagem_servidor,
        linha,
    )

    _bloco(
        "Atendimentos por situação",
        {
            ROTULOS_STATUS.get(status, status): quantidade
            for status, quantidade in contagem_status.items()
        },
        linha,
    )

    _ajustar_larguras(aba, [46, 16, 12])


def _montar_aba_detalhes(
    aba,
    atendimentos: list,
    agregado: bool,
) -> None:
    aba.sheet_view.showGridLines = False

    colunas = ["Senha", "Cidadão", "CPF", "MASP", "Telefone"]

    if agregado:
        # A coluna de setor só faz sentido quando o relatório junta
        # vários setores — em um relatório de setor único ela seria
        # repetitiva (a mesma célula do início ao fim).
        colunas.append("Setor")

    colunas += [
        "Assunto", "Relato do cidadão", "Prioridade", "Situação",
        "Solicitado em", "Convocado em", "Iniciado em", "Finalizado em",
        "Espera (min)", "Duração (min)", "Servidor", "MASP do servidor",
        "Sala", "Resultado", "Observações do atendimento",
    ]

    _escrever_cabecalho(aba, 1, colunas)
    aba.freeze_panes = "A2"

    for indice, atendimento in enumerate(atendimentos, start=2):
        cidadao = atendimento.cidadao

        valores = [
            atendimento.numero_senha or "",
            cidadao.nome if cidadao else "",
            formatar_cpf(cidadao.cpf) if cidadao else "",
            formatar_masp(cidadao.masp) if cidadao else "",
            cidadao.telefone if cidadao else "",
        ]

        if agregado:
            valores.append(
                atendimento.setor.nome if atendimento.setor else ""
            )

        valores += [
            _rotulo_assunto(atendimento.assunto),
            atendimento.descricao or "",
            (
                "Prioritário"
                if atendimento.prioridade == "PRIORITARIO"
                else "Normal"
            ),
            ROTULOS_STATUS.get(atendimento.status, atendimento.status),
            _formatar_datahora(atendimento.data_solicitacao),
            _formatar_datahora(atendimento.data_convocacao),
            _formatar_datahora(atendimento.data_inicio),
            _formatar_datahora(atendimento.data_finalizacao),
            _minutos_entre(
                atendimento.data_solicitacao,
                atendimento.data_inicio or atendimento.data_convocacao,
            ),
            _minutos_entre(
                atendimento.data_inicio,
                atendimento.data_finalizacao,
            ),
            atendimento.servidor_nome or "",
            formatar_masp(atendimento.servidor_masp) or "",
            atendimento.numero_sala or "",
            atendimento.resultado or "",
            atendimento.observacoes or "",
        ]

        for coluna, valor in enumerate(valores, start=1):
            celula = aba.cell(row=indice, column=coluna, value=valor)
            celula.border = BORDA_FINA
            celula.alignment = Alignment(vertical="top", wrap_text=True)

            if indice % 2 == 0:
                celula.fill = PatternFill("solid", fgColor=COR_FAIXA)

    larguras = [10, 30, 18, 14, 16]
    if agregado:
        larguras.append(16)
    larguras += [
        28, 40, 13, 15, 18, 18, 18, 18, 13, 14, 26, 16, 12, 24, 40,
    ]
    _ajustar_larguras(aba, larguras)

    if atendimentos:
        aba.auto_filter.ref = (
            f"A1:{get_column_letter(len(colunas))}{len(atendimentos) + 1}"
        )


def gerar_relatorio(
    atendimentos: list,
    titulo_periodo: str,
    titulo_abrangencia: str,
    agregado: bool = False,
) -> bytes:
    """
    Monta a planilha e devolve os bytes prontos para download.

    ``titulo_periodo``: ex. "Julho de 2026" ou "Semana de 20/07 a
    26/07/2026" — já formatado por quem chama.
    ``titulo_abrangencia``: ex. "Setor A" ou "Todos os setores".
    ``agregado``: True quando o relatório junta mais de um setor — liga
    a coluna/bloco extra de quebra por setor.
    """
    planilha = Workbook()

    aba_resumo = planilha.active
    aba_resumo.title = "Resumo"

    _montar_aba_resumo(
        aba_resumo,
        atendimentos,
        titulo_periodo,
        titulo_abrangencia,
        agregado,
    )

    aba_detalhes = planilha.create_sheet("Atendimentos")
    _montar_aba_detalhes(aba_detalhes, atendimentos, agregado)

    buffer = io.BytesIO()
    planilha.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()
